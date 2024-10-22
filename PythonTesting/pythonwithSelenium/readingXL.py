import openpyxl

#file--> workbook ---> sheets ----> rows--->cells
file = "location of the xl file"
workbook = openpyxl.load_workbook(file)
sheet = workbook["sheet1"]

rows = sheet.max_row # count number of rows in a excel sheet
cols =sheet.max_column # count number of column in a excel sheet

#Reading all the rows & coloumn from excel sheet
for r in range(1,rows + 1):
    for c in range(1, cols +1):
        print(sheet.cell(r,c).value,end=" ")
    print()

#writing data in excel file
#same data in each and every row column
file = "location of the xl file"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active # (or) sheet = workbook["data] ---get active sheet from excel
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c). value = "welcome"
workbook.save(file)

#multiple data
file = "location of the xl file"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

sheet.cell(1,1).value = 123
sheet.cell(1,2).value = "smith"
sheet.cell(1,3).value = "engineer"

sheet.cell(2,1).value = 567
sheet.cell(2,2).value = "john"
sheet.cell(2,3).value = "manager"

sheet.cell(3,1).value = 364
sheet.cell(3,2).value = "david"
sheet.cell(3,3).value = "developer"

workbook.save(file) #save the file after entering the data

