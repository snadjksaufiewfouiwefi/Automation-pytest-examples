import openpyxl
from openpyxl.styles import patternFill

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum,columnno).value = data
    workbook.save(file)

def fillGreencolor(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = patternFill(start_color = '60b212',
                            end_color='60b212',
                            fill_type='soild')
    sheet.cell(rownum,columnno).fill = greenFill
    workbook.save(file)


def fillRedcolor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = patternFill(start_color='ff0000',
                            end_color='ff0000',
                            fill_type='soild')
    sheet.cell(rownum, columnno).fill = redFill
    workbook.save(file)